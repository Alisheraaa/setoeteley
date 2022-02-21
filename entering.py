def worker():
    import openpyxl
    import work_menu
    book = openpyxl.load_workbook("staff.xlsx")
    sheet = book.worksheets[0]
    logins = [sheet[i][2].value for i in range(2, sheet.max_row + 1)]
    nums = [sheet[i][4].value for i in range(2, sheet.max_row + 1)]
    book.close()
    running = True
    push = True
    count = 0
    count1 = 0
    count_tel = 0
    tow = True

    while running:
        login = input('Enter your login:')
        if login in logins:
            while push:
                password = input('Enter your password:')
                if password == sheet[logins.index(login) + 2][3].value:
                    print('Login completed')
                    work_menu.worker_menu(login)

                    running = False
                    push = False
                else:
                    print('Invalid password')
                    count += 1
                    if count == 3:
                        qwest = input('Forgot your password?')
                        if qwest == 'yes':
                            while tow:
                                tell = int(input('Enter your phone number:'))
                                if tell == sheet[logins.index(login) + 2][4].value:

                                    print('Your password:', sheet[logins.index(login) + 2][3].value)
                                    push = False
                                    running = True
                                    tow = False
                                    worker()
                                elif count_tel >= 2:
                                    tow = False
                                    push = False
                                    running = False
                                else:
                                    print('Invalid number')
                                    count_tel += 1
                                    push = False
                                    running = False
                                    continue
                    elif count1 == 5:
                        running = False
                        tow = False
                        push = False



        elif login not in logins:
            count += 1
            if count == 3:
                qw = input('Forgot your login?')
                if qw == "yes":
                    while tow:
                        tell = int(input('Enter your phone number: '))
                        if tell in nums:
                            print('Your login: ', sheet[nums.index(tell) + 2][2].value)
                            push = False
                            running = False
                            tow = False
                            worker()
                        elif count1 == 2:
                            tow = False
                            push = False
                            running = False
                        else:
                            print('Invalid number')
                            count1 +=1
                            push = False
                            running = False
                            continue
            elif count == 5:
                running = False
                tow = False
                push = False

def manager():
    import openpyxl
    import work_menu
    book = openpyxl.load_workbook("staff.xlsx")
    sheet = book.worksheets[1]
    logins = [sheet[i][2].value for i in range(2, sheet.max_row + 1)]
    nums = [sheet[i][4].value for i in range(2, sheet.max_row + 1)]
    book.close()
    running = True
    push = True
    count = 0
    count1 = 0
    count_tel = 0
    tow = True

    while running:
        login = input('Enter your login: ')
        if login in logins:
            while push:
                password = input('Enter your password: ')
                if password == sheet[logins.index(login) + 2][3].value:
                    print('Login completed')
                    work_menu.manager_menu()


                    running = False
                    push = False
                else:
                    print('Invalid password')
                    count += 1
                    if count == 3:
                        qwest = input('Forgot our password? ')
                        if qwest == 'yes':
                            while tow:
                                tell = int(input('Enter your phone number: '))
                                if tell == sheet[logins.index(login) + 2][4].value:

                                    print('Your login: ', sheet[logins.index(login) + 2][2].value)
                                    push = False
                                    running = True
                                    tow = False
                                    manager()
                                elif count_tel >= 2:
                                    tow = False
                                    push = False
                                    running = False
                                else:
                                    print('Invalid number')
                                    count_tel += 1
                                    push = False
                                    running = False
                                    continue
                    elif count1 == 5:
                        running = False
                        tow = False
                        push = False



        elif login not in logins:
            count += 1
            if count == 3:
                qw = input('Forgot your login? ')
                if qw == "yes":
                    while tow:
                        tell = int(input('Enter your phone number: '))
                        if tell in nums:
                            print('Your login: ', sheet[nums.index(tell) + 2][2].value)
                            push = False
                            running = False
                            tow = False
                            manager()
                        elif count1 == 2:
                            tow = False
                            push = False
                            running = False
                        else:
                            print('Invalid number')
                            count1 += 1
                            push = False
                            running = False
                            continue
            elif count == 5:
                running = False
                tow = False
                push = False

def markerting():
    import openpyxl
    import work_menu
    book = openpyxl.load_workbook("staff.xlsx")
    sheet = book.worksheets[2]
    logins = [sheet[i][2].value for i in range(2, sheet.max_row + 1)]
    nums = [sheet[i][4].value for i in range(2, sheet.max_row + 1)]
    book.close()
    running = True
    push = True
    count = 0
    count1 = 0
    count_tel = 0
    tow = True

    while running:
        login = input('Enter your login: ')
        if login in logins:
            while push:
                password = input('Enter your password: ')
                if password == sheet[logins.index(login) + 2][3].value:
                    print('Login completed\n')
                    work_menu.marketing_menu()


                    running = False
                    push = False
                else:
                    count += 1
                    if count == 3:
                        qwest = input('Forgot your password? ')
                        if qwest == 'yes':
                            while tow:
                                tell = int(input('Enter your phone number: '))
                                if tell == sheet[logins.index(login) + 2][4].value:

                                    print('Your login: ', sheet[logins.index(login) + 2][2].value)
                                    push = False
                                    running = True
                                    tow = False
                                    markerting()
                                elif count_tel >= 2:
                                    tow = False
                                    push = False
                                    running = False
                                else:
                                    print('Invalid number')
                                    count_tel += 1
                                    push = False
                                    running = False
                                    continue
                    elif count1 == 5:
                        running = False
                        tow = False
                        push = False



        elif login not in logins:
            count += 1
            if count == 3:
                qw = input('Forgot your login? ')
                if qw == "yes":
                    while tow:
                        tell = int(input('Enter your phone number: '))
                        if tell in nums:
                            print('Your login: ', sheet[nums.index(tell) + 2][2].value)
                            push = False
                            running = False
                            tow = False
                            markerting()
                        elif count1 == 2:
                            tow = False
                            push = False
                            running = False
                        else:
                            print('Invalid number')
                            count1 += 1
                            push = False
                            running = False
                            continue
            elif count == 5:
                running = False
                tow = False
                push = False

def director():
    import openpyxl
    import work_menu
    book = openpyxl.load_workbook("staff.xlsx")
    sheet = book.worksheets[3]
    logins = [sheet[i][2].value for i in range(2, sheet.max_row + 1)]
    nums = [sheet[i][4].value for i in range(2, sheet.max_row + 1)]
    book.close()
    running = True
    push = True
    count = 0
    count1 = 0
    count_tel = 0
    tow = True

    while running:
        login = input('Enter your login: ')
        if login in logins:
            while push:
                password = input('Enter your password: ')
                if password == sheet[logins.index(login) + 2][3].value:
                    print('Login completed')
                    work_menu.director_menu()



                    running = False
                    push = False
                else:
                    print('Invalid password')
                    count += 1
                    if count == 3:
                        qwest = input('Forgot your password? ')
                        if qwest == 'yes':
                            while tow:
                                tell = int(input('Enter your phone number: '))
                                if tell == sheet[logins.index(login) + 2][4].value:

                                    print('Your login: ', sheet[logins.index(login) + 2][2].value)
                                    push = False
                                    running = True
                                    tow = False
                                    director()
                                elif count_tel >= 2:
                                    tow = False
                                    push = False
                                    running = False
                                else:
                                    print('Invalid number')
                                    count_tel += 1
                                    push = False
                                    running = False
                                    continue
                    elif count1 == 5:
                        running = False
                        tow = False
                        push = False



        elif login not in logins:
            count += 1
            if count == 3:
                qw = input('Forgot your login? ')
                if qw == "yes":
                    while tow:
                        tell = int(input('Enter your phone number: '))
                        if tell in nums:
                            print('Your login: ', sheet[nums.index(tell) + 2][2].value)
                            push = False
                            running = False
                            tow = False
                            director()
                        elif count1 == 2:
                            tow = False
                            push = False
                            running = False
                        else:
                            print('Invalid number')
                            count1 += 1
                            push = False
                            running = False
                            continue
            elif count == 5:
                running = False
                tow = False
                push = False

def sales():
    import openpyxl
    import work_menu
    book = openpyxl.load_workbook("staff.xlsx")
    sheet = book.worksheets[4]
    logins = [sheet[i][2].value for i in range(2, sheet.max_row + 1)]
    nums = [sheet[i][4].value for i in range(2, sheet.max_row + 1)]
    book.close()
    running = True
    push = True
    count = 0
    count1 = 0
    count_tel = 0
    tow = True

    while running:
        login = input('Enter your login: ')
        if login in logins:
            while push:
                password = input('Enter your password: ')
                if password == sheet[logins.index(login) + 2][3].value:
                    print('Login completed')
                    work_menu.sales_menu()


                    running = False
                    push = False
                else:
                    print('Invalid password')
                    count += 1
                    if count == 3:
                        qwest = input('Forgot your password? ')
                        if qwest == 'yes':
                            while tow:
                                tell = int(input('Enter your phone number: '))
                                if tell == sheet[logins.index(login) + 2][4].value:

                                    print('Your login: ', sheet[logins.index(login) + 2][2].value)
                                    push = False
                                    running = True
                                    tow = False
                                    sales()
                                elif count_tel >= 2:
                                    tow = False
                                    push = False
                                    running = False
                                else:
                                    print('Invalid number')
                                    count_tel += 1
                                    push = False
                                    running = False
                                    continue
                    elif count1 == 5:
                        running = False
                        tow = False
                        push = False



        elif login not in logins:
            count += 1
            if count == 3:
                qw = input('Forgot your login? ')
                if qw == "yes":
                    while tow:
                        tell = int(input('Enter your phone number: '))
                        if tell in nums:
                            print('Your login: ', sheet[nums.index(tell) + 2][2].value)
                            push = False
                            running = False
                            tow = False
                            sales()
                        elif count1 == 2:
                            tow = False
                            push = False
                            running = False
                        else:
                            print('Invalid number')
                            count1 += 1
                            push = False
                            running = False
                            continue
            elif count == 5:
                running = False
                tow = False
                push = False