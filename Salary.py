import openpyxl
book = openpyxl.load_workbook('Zones.xlsx')
sheet = book.active
book.close()
workbook_name = 'staff.xlsx'
book_2 = openpyxl.load_workbook(workbook_name)
sheet_2 = book_2.worksheets[5]


names = [sheet_2[i][0].value for i in range(2,sheet_2.max_row)]

def up():
    running = True
    push = True
    while running:
        name = input('Enter the name of the worker to whom you want to raise the salary:')
        if name in names:
            while push:
                zp = int(input('Enter the amount of the salary supplement: '))
                sheet_2[names.index(name)+2][2].value += zp
                book_2.save(filename=workbook_name)
                print(f'The surcharge has been successfully made! The current salary for this worker after the allowance: {sheet_2[names.index(name)+2][2].value}')
                running = False
                push = False
        else:
            running = False
            print('Worker not found')
            continue

def down():
    running = True
    push = True
    while running:
        name = input('Type the name of the worker to whom you want to lower the salary:')
        if name in names:
            while push:
                zp = int(input('Enter the deduction amount: '))
                sheet_2[names.index(name) + 2][2].value -= zp
                print(f'The deduction has been successfully made! The current salary for this worker after deduction: {sheet_2[names.index(name) + 2][2].value}')
                book_2.save(filename=workbook_name)
                running = False
                push = False
        else:
            running = False
            print('Worker not found')
            continue



def delete():
    running = True
    push = True
    while running:
        name = input('Type the name of the worker you want to delete:')
        if name in names:
            while push:
                del name
                print('Worker successfully deleted')
                running = False
                push = False
        else:
            running = False
            print('Worker not found')
            continue

# def delete():
#     running = True
#     push = True
#     while running:
#         name = input('Type the name of the worker you want to delete:')
#         if name in names:
#             while push:
#                 sheet_2.delete_rows[names.index(name)].value
#                 print('Worker successfully deleted')
#                 running = False
#                 push = False
#         else:
#             running = False
#             print('Worker not found')
#             continue


