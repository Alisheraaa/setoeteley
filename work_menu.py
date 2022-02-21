import Salary


def worker_menu(login):
    import openpyxl
    book = openpyxl.load_workbook('Tasks.xlsx')
    sheet = book.active
    workbook_name = 'staff.xlsx'
    book_2 = openpyxl.load_workbook(workbook_name)
    sheet_2 = book_2.worksheets[5]
    book_3 = openpyxl.load_workbook('Tasks.xlsx')
    sheet_3 = book_3.worksheets[0]
    logins = [sheet_3[i][4].value for i in range(2, sheet_3.max_row + 1)]
    print(f'Welcome dear Worker!\n')
    print('Please dial the menu number to work with the program, if finished, then dial 5:\n')
    print('1.Show the list of tasks assigned to me.')
    print('2.Show a list of completed instructions.')
    print('3.Show salary')
    print('4.Exit\n')
    running = True
    while running:
        answer = int(input(':'))
        if answer == 1:
            print(f'Your task: {sheet_3[logins.index(login) + 2][2].value}')
            done = input('Have you completed the task? ')
            if done == "yes":
                if sheet_3[logins.index(login) + 2][5].value != None:
                    sheet_3[logins.index(login) + 2][5].value += '/' + sheet_3[logins.index(login) + 2][2].value
                    sheet_3[logins.index(login) + 2][2].value = ''
                    book_3.save(filename="Tasks.xlsx")
                    print('The task is recorded!')
                elif sheet_3[logins.index(login) + 2][5].value == None:
                    sheet_3[logins.index(login) + 2][5].value = sheet_3[logins.index(login) + 2][2].value
                    sheet_3[logins.index(login) + 2][2].value = ''
                    book_3.save(filename="Tasks.xlsx")
                    print('The task is recorded!')
            else:
                continue


        elif answer == 2:
            print(sheet_3[logins.index(login) + 2][5].value)
        elif answer == 3:
            print(f'Your salary: {sheet_2[logins.index(login) + 2][2].value}')
        elif answer == 4:
            running = False
            print('The program is completed, we will be glad to see you back!')


def marketing_menu():
    import openpyxl
    book = openpyxl.load_workbook('zones.xlsx')
    sheet = book.active
    workbook_name = 'staff.xlsx'
    book_2 = openpyxl.load_workbook(workbook_name)
    sheet_2 = book_2.worksheets[5]

    running = True
    tow = True
    push = True

    print(f'Welcome dear Marketing specialist!\n')
    print('Please dial the menu number to work with the program, if finished, then dial 6:\n')
    print('1.Show a list of all clients coverage areas by region')
    print('2.Show a list of categories for marketing')
    print('3.Show the allocated budget for a certain category of places for marketing')
    print('4.Show the total budget for marketing')
    print('5.Spend the budget on promotion')
    print('6.Exit')
    while running:
        answer = int(input(':'))
        if answer == 1:
            cities = [sheet[i][1].value for i in range(2, sheet.max_row)]
            sum_ = sum(cities)
            print(f'List of all coverage areas: ')
            print(f"Bishkek - number of clients: {sheet[2][1].value}-{round(sheet[2][1].value * 100 / sum_)}%")
            print(f"Osh - number of clients: {sheet[3][1].value}-{round(sheet[3][1].value * 100 / sum_)}%")
            print(f"Jalal-abad- number of clients: {sheet[4][1].value}-{round(sheet[4][1].value * 100 / sum_)}%")
            print(f"Talas - number of clients: {sheet[5][1].value}-{round(sheet[5][1].value * 100 / sum_)}%")
            print(f"Naryn - number of clients: {sheet[6][1].value}-{round(sheet[6][1].value * 100 / sum_)}%")
            print(f"Issyk-Kul - number of clients: {sheet[7][1].value}-{round(sheet[7][1].value * 100 / sum_)}%")
            print(f"Batken - number of clients: {sheet[8][1].value}-{round(sheet[8][1].value * 100 / sum_)}%")
        elif answer == 2:
            print("List of categories for marketing: ")
            print(f"Instagram - {sheet[2][4].value} users")
            print(f"Facebook - {sheet[3][4].value} users")
            print(f"VK - {sheet[4][4].value} users")
            print(f"Twitter - {sheet[5][4].value} users")
            print(f"Telegram - {sheet[6][4].value} users")
            print(f"2GIS - {sheet[7][4].value} users")
        elif answer == 3:
            print(f'Dedicated budget for Instagram: {sheet_2[2][5].value}')
            print(f'Dedicated budget for Facebook: {sheet_2[3][5].value}')
            print(f'Dedicated budget for VK: {sheet_2[4][5].value}')
            print(f'Dedicated budget for Twitter: {sheet_2[5][5].value}')
            print(f'Dedicated budget for Youtube: {sheet_2[6][5].value}')
            print(f'Dedicated budget for 2GIS: {sheet_2[7][5].value}')
        elif answer == 4:
            print(f"Marketing budget: {sheet_2['G2'].value} ")
            print(f"Current means for marketing: {sum([sheet_2[i][5].value for i in range(2, 8)])}")
        elif answer == 5:
            print('1.Instagram')
            print('2.Facebook')
            print('3.VK')
            print('4.Twitter')
            print('5.Youtube')
            print('6.2GIS\n')
            print('Enter 0 to cancel')
            while push:
                spend_sub = int(input('Choose a name to promote: '))
                if spend_sub == 1:
                    while tow:
                        summa = int(input('Enter the amount of expenditure that you want to spend from the budget: '))
                        if summa <= sheet_2[2][5].value:
                            sheet_2[2][5].value -= summa
                            print('Operation completed successfully')
                            book_2.save(filename=workbook_name)
                            tow = False
                            push = False
                        elif summa > sheet_2[2][5].value:
                            print('Value exceeds budget!')
                            continue

                elif spend_sub == 2:
                    while tow:
                        summa = int(input('Enter the amount of expenditure that you want to spend from the budget: '))
                        if summa <= sheet_2[3][5].value:
                            sheet_2[3][5].value -= summa
                            print('Operation completed successfully')
                            book_2.save(filename=workbook_name)
                            # book_2.close()
                            tow = False
                            push = False
                        elif summa == 0:
                            tow = False
                        elif summa > sheet_2[3][5].value:
                            print('Value exceeds budget!')
                            continue
                elif spend_sub == 3:
                    summa = int(input('Enter the amount of expenditure that you want to spend from the budget: '))
                    if summa <= sheet_2[4][5].value:
                        sheet_2[4][5].value -= summa
                        print('Operation completed successfully')
                        book_2.save(filename=workbook_name)
                        # book_2.close()
                        tow = False
                        push = False
                    elif summa == 0:
                        tow = False
                    elif summa > sheet_2[4][5].value:
                        print('Value exceeds budget!')
                        continue

                elif spend_sub == 4:
                    summa = int(input('Enter the amount of expenditure that you want to spend from the budget: '))
                    if summa <= sheet_2[5][5].value:
                        sheet_2[5][5].value -= summa
                        print('Operation completed successfully')
                        book_2.save(filename=workbook_name)
                        # book_2.close()
                        tow = False
                        push = False
                    elif summa == 0:
                        tow = False
                    elif summa > sheet_2[5][5].value:
                        print('Value exceeds budget!')
                        continue
                elif spend_sub == 5:
                    summa = int(input('Enter the amount of expenditure that you want to spend from the budget: '))
                    if summa <= sheet_2[6][5].value:
                        sheet_2[6][5].value -= summa
                        print('Operation completed successfully')
                        book_2.save(filename=workbook_name)
                        # book_2.close()
                        tow = False
                        push = False
                    elif summa == 0:
                        tow = False
                    elif summa > sheet_2[6][5].value:
                        print('Value exceeds budget!')
                        continue
                elif spend_sub == 6:
                    summa = int(input('Enter the amount of expenditure that you want to spend from the budget: '))
                    if summa <= sheet_2[7][5].value:
                        sheet_2[7][5].value -= summa
                        print('Operation completed successfully')
                        book_2.save(filename=workbook_name)
                        # book_2.close()
                        tow = False
                        push = False
        elif answer == 6:
            print('The program is completed, we will be glad to see you back!')
            break


def director_menu():
    import openpyxl
    book = openpyxl.load_workbook('zones.xlsx')
    sheet = book.active
    book.close()
    book_2 = openpyxl.load_workbook('staff.xlsx')
    sheet_2 = book_2.worksheets[5]

    running = True
    print('Welcome dear Director!\n')
    print('Please dial the menu number to work with the program, if finished, then dial 9: ')
    print('1.Show a list of all coverage areas ')
    print('2.Show a list of budget categories ')
    print('3.Show the allocated budget for a certain category of places for marketing ')
    print('4.Show current marketing funds')
    print('5.Show the total budget required for the salary')
    print("6.Increase employee's salary")
    print("7.Lower an employee's salary")
    print('8.Show a list of equipment for the construction of objects ')
    print('9.Dismiss an employee')
    print('10.Exit')
    while running:
        answer = int(input(':'))
        if answer == 1:
            cities = [sheet[i][1].value for i in range(2, sheet.max_row)]
            sum_ = sum(cities)
            print(f'List of all coverage areas: ')
            print(f"Bishkek - number of clients: {sheet[2][1].value}-{round(sheet[2][1].value * 100 / sum_)}%")
            print(f"Osh - number of clients: {sheet[3][1].value}-{round(sheet[3][1].value * 100 / sum_)}%")
            print(f"Jalal-abad - number of clients: {sheet[4][1].value}-{round(sheet[4][1].value * 100 / sum_)}%")
            print(f"Talas - number of clients: {sheet[5][1].value}-{round(sheet[5][1].value * 100 / sum_)}%")
            print(f"Naryn - number of clients: {sheet[6][1].value}-{round(sheet[6][1].value * 100 / sum_)}%")
            print(f"Issyk-Kul - number of clients: {sheet[7][1].value}-{round(sheet[7][1].value * 100 / sum_)}%")
            print(f"Batken - number of clients: {sheet[8][1].value}-{round(sheet[8][1].value * 100 / sum_)}%")

        elif answer == 2:
            print(f"Budget for salaries: {sum([sheet_2[i][2].value for i in range(2, sheet_2.max_row)])}")
            print(f"Budget for marketing: {sheet_2['G2'].value} ")
        elif answer == 3:
            print(f'Allocated budget for Instagram: {sheet_2[2][5].value}')
            print(f'Allocated budget for Facebook: {sheet_2[3][5].value}')
            print(f'Allocated budget for VK: {sheet_2[4][5].value}')
            print(f'Allocated budget for Twitter: {sheet_2[5][5].value}')
            print(f'Allocated budget for Telegram: {sheet_2[6][5].value}')
            print(f'Allocated budget for 2GIS: {sheet_2[7][5].value}')
        elif answer == 4:
            print(f"Current means for marketing: {sum([sheet_2[i][5].value for i in range(2, 8)])}")
        elif answer == 5:
            book_3 = openpyxl.load_workbook('staff.xlsx')
            sheet_3 = book_3.worksheets[5]
            print(f"The total budget required for salaries: {sum([sheet_3[i][2].value for i in range(2, sheet_2.max_row)])}")
        elif answer == 6:
            Salary.up()
        elif answer == 7:
            Salary.down()
        elif answer == 8:
            sheet_4 = book_2.worksheets[6]
            things = [sheet_4[i][0].value for i in range(2, sheet_4.max_row + 1)]
            items = [sheet_4[i][1].value for i in range(2, sheet_4.max_row + 1)]
            for i, j in zip(things, items):
                print(i + '-' + str(j) + "шт")
        elif answer == 9:
            Salary.delete()
        elif answer == 10:
            print('The program is completed, we will be glad to see you back!')
            running = False
        else:
            continue


def manager_menu():
    import openpyxl
    book_2 = openpyxl.load_workbook('staff.xlsx')
    sheet_2 = book_2.worksheets[5]
    book_3 = openpyxl.load_workbook('Tasks.xlsx')
    sheet_3 = book_3.worksheets[0]
    workers = [sheet_3[i][0].value for i in range(2, sheet_3.max_row + 1)]
    jobs = [sheet_3[i][3].value for i in range(2, sheet_3.max_row) if sheet_3[i][3].value != None]
    free_workers = [sheet_3[i][0].value for i in range(2, sheet_3.max_row + 1) if sheet_3[i][2].value == None]

    running = True
    tow = True
    push = True

    print('Welcome dear Manager!\n')
    print('Please dial the menu number to work with the program, if finished, then dial 5:')
    print("1.Show the list of workers ")
    print("2.Show the list of tasks ")
    print('3.Add the task to a worker')
    print('4.Add to the list of tasks')
    print("5.Show a list of instructions to workers")
    print("6.Show a list of all coverage areas ")
    print("7.Exit ")
    while running:
        answer = int(input(":"))
        if answer == 1:
            print('List of workers: ')
            name = [sheet_2[i][0].value for i in range(2, sheet_2.max_row)]
            job = [sheet_2[i][1].value for i in range(2, sheet_2.max_row)]
            for i, j in zip(name, job):
                print(i + '-' + str(j))
        elif answer == 2:
            book = openpyxl.load_workbook('Tasks.xlsx')
            sheet = book.worksheets[0]
            print('To-do list: ')
            jobs = [sheet[i][3].value for i in range(2, sheet.max_row) if sheet[i][3].value != None]
            for i in jobs:
                print(i)
        elif answer == 3:
            if jobs == []:
                print('No tasks')
                continue
            else:
                print('List of available workers: ')
                print(free_workers)
                while push:
                    free_worker_name = input('Enter the name of the worker to whom you want to add the task: ')
                    if free_worker_name in free_workers:
                        while tow:
                            for i, item in enumerate(jobs):
                                print(i + 1, item)
                            choose_job = input('Enter a task for this worker:')
                            if choose_job in jobs:
                                for i in range(1, sheet_3.max_row + 1):
                                    if sheet_3[i][3].value == choose_job:
                                        sheet_3[workers.index(free_worker_name) + 2][2].value = sheet_3[i][3].value
                                        sheet_3[i][3].value = ''
                                        book_3.save(filename='Tasks.xlsx')
                                        print('Operation completed successfully!')
                                        tow = False
                                        push = False
        elif answer == 4:
            task = input('Enter the task: ')
            for i in range(2, sheet_3.max_row + 1):
                if sheet_3[i][3].value == None:
                    sheet_3[i][3].value = task
                    book_3.save(filename='Tasks.xlsx')
                    print('Operation completed successfully')
                    push = False
                    tow = False
                    break


        elif answer == 5:
            process = jobs = [sheet_3[i][2].value for i in range(2, sheet_3.max_row)]
            for i, j in zip(workers, process):
                print(f'{i} - {j}')
        elif answer == 6:
            book = openpyxl.load_workbook('zones.xlsx')
            sheet = book.active
            workbook_name = 'staff.xlsx'
            book_2 = openpyxl.load_workbook(workbook_name)
            sheet_2 = book_2.worksheets[5]
            book_3 = openpyxl.load_workbook('staff.xlsx')
            sheet_3 = book_3.worksheets[5]

            running = True
            tow = True
            push = True

            cities = [sheet[i][1].value for i in range(2, sheet.max_row)]
            sum_ = sum(cities)
            print(f'List of all coverage areas: ')
            print(f"Bishkek - number of clients: {sheet[2][1].value}-{round(sheet[2][1].value * 100 / sum_)}%")
            print(f"Osh - number of clients: {sheet[3][1].value}-{round(sheet[3][1].value * 100 / sum_)}%")
            print(f"Jalal-Abad - number of clients: {sheet[4][1].value}-{round(sheet[4][1].value * 100 / sum_)}%")
            print(f"Talas - number of clients: {sheet[5][1].value}-{round(sheet[5][1].value * 100 / sum_)}%")
            print(f'Naryn - number of clients: {sheet[6][1].value}-{round(sheet[6][1].value * 100 / sum_)}%')
            print(f"Issyk-Kul - number of clients: {sheet[7][1].value}-{round(sheet[7][1].value * 100 / sum_)}%")
            print(f"Batken - number of clients: {sheet[8][1].value}-{round(sheet[8][1].value * 100 / sum_)}%")
        elif answer == 7:
            print('\nThe program is completed, we will be glad to see you back!')
            running = False
            tow = False
            push = False


def sales_menu():
    import openpyxl
    book_3 = openpyxl.load_workbook('Tasks.xlsx')
    sheet_3 = book_3.worksheets[1]
    clients = [sheet_3[i][1].value for i in range(2, sheet_3.max_row + 1) if sheet_3[i][1].value != None]
    rooms = [sheet_3[i][0].value for i in range(2, sheet_3.max_row + 1) if sheet_3[i][0].value != None]
    all = [sheet_3[i][1].value for i in range(2, sheet_3.max_row + 1)]
    free_rooms = [sheet_3[i][0].value for i in range(2, sheet_3.max_row + 1) if sheet_3[i][1].value == None]
    running = True

    print('Welcome dear Sale-manager!\nPlease dial the menu number to work with the program, if finished, then dial 6:\n')
    print('1.Show all clients ')
    print('2.Clients Search')
    print('3.Show available rooms')
    print('4.Show the most expensive room')
    print('5.Show the cheapest room')
    print('6.Exit')
    while running:
        answer = int(input(':'))
        if answer == 1:
            print('List of all clients: ')
            clients = [sheet_3[i][1].value for i in range(2, sheet_3.max_row + 1) if sheet_3[i][1].value != None]
            for i in clients:
                print(i)
        elif answer == 2:
            name = input("Enter the client's name:")
            if name in clients:
                print(name)
                print('Room №', sheet_3[all.index(name) + 2][0].value)
            else:
                print('Client not found')
                continue
        elif answer == 4:
            print(f'The most expensive room: {max([sheet_3[j][2].value for j in range(2, sheet_3.max_row + 1)])}')
            s = [sheet_3[j][2].value for j in range(2, sheet_3.max_row + 1)]
            id = s.index(max(s))
            print(f'Room №{sheet_3[id + 2][0].value}')
        elif answer == 3:
            print('Free rooms: ')
            for free in free_rooms:
                print('№', free)
        elif answer == 5:
            print(f'The cheapest room: {min([sheet_3[j][2].value for j in range(2, sheet_3.max_row + 1)])}')
            s = [sheet_3[j][2].value for j in range(2, sheet_3.max_row + 1)]
            id = s.index(min(s))
            print(f'Room №{sheet_3[id + 2][0].value}')
        elif answer == 6:
            print('The program is completed, we will be glad to see you back')
            running = False
